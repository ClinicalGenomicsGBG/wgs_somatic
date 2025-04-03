from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.utils.cell import coordinate_from_string as cfs
import pandas as pd
import glob


def process_qc_files(file_list, output_file):
    """
    Processes a list of QC stats files and outputs a combined Excel file.

    Args:
        file_list (list): List of file paths to process.
        output_file (str): Path to the output Excel file.
    """
    def convert_rng_to_df(tlc, l_col, l_row, sheet):
        first_col = cfs(tlc)[0]
        first_row = cfs(tlc)[1]

        # Define the range for the data (including the header row)
        rng = f"{first_col}{first_row}:{l_col}{l_row}"

        data_rows = []
        for row in sheet[rng]:
            data_rows.append([cell.value for cell in row])

        # Use the first row as column headers and the rest as data
        return pd.DataFrame(data_rows[1:], columns=data_rows[0])

    section_headers = ['COVERAGE STATS', "TUMOR/NORMAL MATCH-CHECK"]

    all_coverage_tables = []
    all_match_check_tables = []

    for filename in file_list:
        wb = load_workbook(filename)
        ws = wb['qc_stats']

        df_dict = {}
        for cell in ws['A']:
            if cell.value in section_headers:
                tblname = cell.value
                tlc = cell.offset(row=1, column=0).coordinate
                start_row = cfs(tlc)[1] + 1
                for x in range(1, ws.max_column + 1):
                    if cell.offset(row=2, column=x).value is None:
                        last_col = gcl(x)
                        break
                for y in range(2, ws.max_row):
                    if cell.offset(row=y, column=0).value is None:
                        last_row = (start_row + y) - 2
                        break

                df_dict[tblname] = convert_rng_to_df(tlc, last_col, last_row, ws)

        if 'COVERAGE STATS' in df_dict:
            coverage_stats = df_dict['COVERAGE STATS']
            if {'Samplename', 'MEAN_COVERAGE', 'SD_COVERAGE', 'PCT_10X'}.issubset(coverage_stats.columns):
                coverage_table = coverage_stats[['Samplename', 'MEAN_COVERAGE', 'SD_COVERAGE', 'PCT_10X']]
                coverage_table = coverage_table[coverage_table['Samplename'].notna()]
                all_coverage_tables.append(coverage_table)

        if 'TUMOR/NORMAL MATCH-CHECK' in df_dict:
            tumor_normal_match_check = df_dict['TUMOR/NORMAL MATCH-CHECK']
            if {'match_fraction', 'match_status'}.issubset(tumor_normal_match_check.columns):
                match_check_subset = tumor_normal_match_check[['match_fraction', 'match_status']]
                all_match_check_tables.append(match_check_subset)

    # Combine all tables
    if all_coverage_tables:
        combined_coverage_table = pd.concat(all_coverage_tables, ignore_index=True)
    else:
        combined_coverage_table = pd.DataFrame()

    if all_match_check_tables:
        combined_match_check_table = pd.concat(all_match_check_tables, ignore_index=True)
    else:
        combined_match_check_table = pd.DataFrame()

    # Combine the coverage tables and match check tables into one overall table
    if not combined_coverage_table.empty and not combined_match_check_table.empty:
        combined_coverage_table.reset_index(drop=True, inplace=True)
        combined_match_check_table.reset_index(drop=True, inplace=True)

        if len(combined_coverage_table) == len(combined_match_check_table):
            overall_table = pd.concat([combined_coverage_table, combined_match_check_table], axis=1)
        else:
            raise ValueError("Mismatch in the number of rows between coverage and match check tables. Cannot combine.")
    else:
        raise ValueError("One or both of the tables are missing. Cannot create an overall table.")

    # Save the combined table to an Excel file
    overall_table.to_excel(output_file, index=False)
    print(f"Combined table saved to {output_file}")


# Example usage:
file_pattern = '/oldseqstore/workspace/carolina/wgs_somatic/tools/*_qc_stats.xlsx'
file_list = glob.glob(file_pattern)
output_file = '/oldseqstore/workspace/carolina/wgs_somatic/tools/combined_qc_stats.xlsx'

process_qc_files(file_list, output_file)